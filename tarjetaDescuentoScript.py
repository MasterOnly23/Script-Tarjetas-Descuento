from decouple import config
import os
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

date_str = datetime.now().strftime('%d-%m-%Y')

file_path = config('FILE_PATH') + ' ' + date_str + config('FILE_EXTENSION')

#cargar workbook
wb = openpyxl.load_workbook(file_path)

#cargar hoja
sheet = wb[config('SHEET1')]

# Obtener los valores tarjeta Descuento
valores_tarjeta_descuento = [celda.value for celda in sheet['F'][1:]]

#valores duplicados
duplicados = set()
unicos = set()
for valor in valores_tarjeta_descuento:
    if valor in unicos:
        unicos.remove(valor)
        duplicados.add(valor)
    else:
        unicos.add(valor)


# Aplicar un filtro de color a las celdas 
fill = PatternFill(start_color="FF8181", end_color="FF8181", fill_type="solid")
for celda in sheet['F'][1:]:
    if celda.value in duplicados:
        celda.fill = fill


sheet2 = wb.create_sheet(config('SHEET2'))

encabezados = ["Cod_sucursal", "Sucursal", "tarjeta Descuento", "Item Total"]
for i, encabezado in enumerate(encabezados, start=1):
    sheet2.cell(row=1, column=i, value=encabezado)
    

# Copiar valores de la hoja1 a la hoja2 solo si el valor de "tarjeta Descuento" está en duplicados
fila_destino_index = 2
for fila_origen in sheet.iter_rows(min_row=2):
    tarjeta_descuento = fila_origen[5].value  # Obtener valor de "tarjeta Descuento" de la fila actual

    if tarjeta_descuento in duplicados:
        # Copiar solo las columnas especificadas a la hoja2
        datos_copiar = [fila_origen[0].value, fila_origen[1].value, tarjeta_descuento, fila_origen[10].value] 
        for columna, valor in enumerate(datos_copiar, start=1):
            sheet2.cell(row=fila_destino_index, column=columna, value=valor)
        fila_destino_index += 1


sheet3 = wb.create_sheet(config('SHEET3'))
encabezados3 = ["tarjeta Descuento", "Cantidad", "Importe con descuento", "Importe sin descuento", "Descuento aplicado"]
for i, encabezado in enumerate(encabezados3, start=1):
    sheet3.cell(row=1, column=i, value=encabezado)

fila_destino_index_sheet3 = 2
for fila_origen in sheet.iter_rows(min_row=2):
    tarjeta_descuento = fila_origen[5].value  # Obtener valor de "tarjeta Descuento" de la fila actual

    if tarjeta_descuento in duplicados:
        # Copiar solo las columnas especificadas a la hoja2
        datos_copiar_sheet3 = [tarjeta_descuento, fila_origen[8].value] 
        for columna, valor in enumerate(datos_copiar_sheet3, start=1):
            sheet3.cell(row=fila_destino_index_sheet3, column=columna, value=valor)
        fila_destino_index_sheet3 += 1


# Fórmula para "Importe con descuento"
for fila in range(2, fila_destino_index_sheet3):
    celda_importe_con_descuento = sheet3.cell(row=fila, column=3)
    formula = "=SUMIF('Ventas tarjeta sucursal'!C:C, 'Usos importe descuento'!A:A, 'Ventas tarjeta sucursal'!D:D)"
    celda_importe_con_descuento.value = formula
    celda_importe_con_descuento.number_format = '#.##'

# Fórmula para "Importe sin descuento"
for fila in range(2, fila_destino_index_sheet3):
    celda_importe_sin_descuento = sheet3.cell(row=fila, column=4)
    celda_importe_sin_descuento.value = f'=C{celda_importe_sin_descuento.row}/0.9'
    celda_importe_sin_descuento.number_format = '#.##'

# Fórmula para "Descuento aplicado"
for fila in range(2, fila_destino_index_sheet3):
    celda_descuento = sheet3.cell(row=fila, column=5)
    celda_importe_con_descuento = sheet3.cell(row=fila, column=3)
    celda_importe_sin_descuento = sheet3.cell(row=fila, column=4)
    celda_descuento.value = f'={celda_importe_con_descuento.coordinate}-{celda_importe_sin_descuento.coordinate}'
    celda_descuento.number_format = '#.##'


#Ajustar el ancho de las columnas
for sheet in [sheet, sheet2, sheet3]:
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column if cell.value]
        if column:
            max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width


#Filtros
sheet.auto_filter.ref = sheet.dimensions
sheet2.auto_filter.ref = sheet2.dimensions
sheet3.auto_filter.ref = sheet3.dimensions


#Save workbook  
filename_save = date_str + ' ' + config('FILENAME_SAVE')

file_path_processed = os.path.join(config('FILE_PATH_PROCESSED'), filename_save)
wb.save(file_path_processed)



print("Proceso terminado")